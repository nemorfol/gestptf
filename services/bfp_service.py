from database import get_db


BFP_FIELDS = [
    "tipologia", "serie", "data_sottoscrizione", "scadenza",
    "valore_nominale", "valore_rimborso_lordo", "valore_lordo_attuale",
    "ritenuta_fiscale", "valore_rimborso_netto", "valore_lordo_scadenza",
    "ritenuta_scadenza", "valore_netto_scadenza", "regolato_su", "note",
]


def get_all_bfp():
    """Returns all BFP records ordered by data_sottoscrizione DESC."""
    try:
        db = get_db()
        rows = db.execute(
            "SELECT * FROM bfp ORDER BY data_sottoscrizione DESC"
        ).fetchall()
        db.close()
        return [dict(r) for r in rows]
    except Exception as e:
        return {"error": str(e)}


def get_bfp_by_id(id):
    """Returns a single BFP record by id."""
    try:
        db = get_db()
        row = db.execute("SELECT * FROM bfp WHERE id = ?", (id,)).fetchone()
        db.close()
        if row is None:
            return {"error": "Record non trovato"}
        return dict(row)
    except Exception as e:
        return {"error": str(e)}


def create_bfp(data):
    """Insert a new BFP record."""
    try:
        db = get_db()
        placeholders = ", ".join(["?"] * len(BFP_FIELDS))
        columns = ", ".join(BFP_FIELDS)
        values = [data.get(f) for f in BFP_FIELDS]

        cursor = db.execute(
            f"INSERT INTO bfp ({columns}) VALUES ({placeholders})",
            values,
        )
        db.commit()
        new_id = cursor.lastrowid
        db.close()
        return {"id": new_id}
    except Exception as e:
        return {"error": str(e)}


def update_bfp(id, data):
    """Update an existing BFP record."""
    try:
        db = get_db()
        set_clause = ", ".join([f"{f} = ?" for f in BFP_FIELDS])
        values = [data.get(f) for f in BFP_FIELDS]
        values.append(id)

        db.execute(
            f"UPDATE bfp SET {set_clause} WHERE id = ?",
            values,
        )
        db.commit()
        db.close()
        return {"success": True}
    except Exception as e:
        return {"error": str(e)}


def delete_bfp(id):
    """Delete a BFP record by id."""
    try:
        db = get_db()
        db.execute("DELETE FROM bfp WHERE id = ?", (id,))
        db.commit()
        db.close()
        return {"success": True}
    except Exception as e:
        return {"error": str(e)}


def get_bfp_riepilogo():
    """Returns BFP summary grouped by tipologia."""
    try:
        db = get_db()
        rows = db.execute(
            """SELECT tipologia,
                      COUNT(*) as count,
                      SUM(valore_nominale) as totale_nominale,
                      SUM(valore_lordo_attuale) as totale_lordo_attuale,
                      SUM(valore_netto_scadenza) as totale_netto_scadenza
               FROM bfp
               GROUP BY tipologia"""
        ).fetchall()
        db.close()
        return [dict(r) for r in rows]
    except Exception as e:
        return {"error": str(e)}


def get_bfp_summary():
    """Returns overall BFP totals."""
    try:
        db = get_db()
        row = db.execute(
            """SELECT SUM(valore_nominale) as totale_nominale,
                      SUM(valore_lordo_attuale) as totale_attuale,
                      SUM(valore_rimborso_netto) as totale_rimborso_netto,
                      SUM(valore_netto_scadenza) as totale_netto_scadenza,
                      COUNT(*) as totale_count
               FROM bfp"""
        ).fetchone()
        db.close()
        return dict(row)
    except Exception as e:
        return {"error": str(e)}


def _parse_euro_value(value):
    """Parse Italian euro format (e.g. '€1.234,56' or '1.234,56 €') to float."""
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    s = s.replace("€", "").replace("\u20ac", "").strip()
    # Italian format: dots as thousands separator, comma as decimal
    s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def _parse_date(value):
    """Parse date from dd/mm/yyyy or '30 mar 2046' style to yyyy-mm-dd string."""
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None

    # Try dd/mm/yyyy
    import re
    match = re.match(r"(\d{1,2})/(\d{1,2})/(\d{4})", s)
    if match:
        day, month, year = match.groups()
        return f"{year}-{month.zfill(2)}-{day.zfill(2)}"

    # Try "30 mar 2046" style
    mesi = {
        "gen": "01", "feb": "02", "mar": "03", "apr": "04",
        "mag": "05", "giu": "06", "lug": "07", "ago": "08",
        "set": "09", "ott": "10", "nov": "11", "dic": "12",
        "jan": "01", "feb": "02", "mar": "03", "apr": "04",
        "may": "05", "jun": "06", "jul": "07", "aug": "08",
        "sep": "09", "oct": "10", "nov": "11", "dec": "12",
    }
    match = re.match(r"(\d{1,2})\s+(\w{3})\s+(\d{4})", s)
    if match:
        day, month_str, year = match.groups()
        month_num = mesi.get(month_str.lower())
        if month_num:
            return f"{year}-{month_num}-{day.zfill(2)}"

    return s


# Mapping from Excel headers to DB columns
_HEADER_MAP = {
    "SCADENZA": "scadenza",
    "TIPOLOGIA": "tipologia",
    "VALORE DI RIMBORSO LORDO": "valore_rimborso_lordo",
    "VALORE LORDO ATTUALE": "valore_lordo_attuale",
    "RITENUTA FISCALE": "ritenuta_fiscale",
    "VALORE RIMBORSO NETTO": "valore_rimborso_netto",
    "VALORE LORDO A SCADENZA": "valore_lordo_scadenza",
    "RITENUTA FISCALE A SCADENZA": "ritenuta_scadenza",
    "VALORE NETTO A SCADENZA": "valore_netto_scadenza",
    "DATA SOTTOSCRIZIONE": "data_sottoscrizione",
    "VALORE NOMINALE": "valore_nominale",
    "SERIE": "serie",
    "REGOLATO SU": "regolato_su",
}

_EURO_COLUMNS = {
    "valore_rimborso_lordo", "valore_lordo_attuale", "ritenuta_fiscale",
    "valore_rimborso_netto", "valore_lordo_scadenza", "ritenuta_scadenza",
    "valore_netto_scadenza", "valore_nominale",
}

_DATE_COLUMNS = {"scadenza", "data_sottoscrizione"}


def import_bfp_from_excel(filepath):
    """Import BFP data from an RPOL Excel file using openpyxl."""
    try:
        import openpyxl

        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active

        # Find header row
        headers = []
        header_row = None
        for row in ws.iter_rows(min_row=1, max_row=20):
            cell_values = [str(c.value).strip() if c.value else "" for c in row]
            if "TIPOLOGIA" in cell_values:
                headers = cell_values
                header_row = row[0].row
                break

        if not headers:
            return {"error": "Intestazioni non trovate nel file Excel"}

        # Map column indices to DB field names
        col_map = {}
        for idx, header in enumerate(headers):
            header_upper = header.upper().strip()
            if header_upper in _HEADER_MAP:
                col_map[idx] = _HEADER_MAP[header_upper]

        if not col_map:
            return {"error": "Nessuna colonna riconosciuta nel file Excel"}

        db = get_db()
        count = 0

        for row in ws.iter_rows(min_row=header_row + 1):
            cell_values = [c.value for c in row]
            # Skip empty rows
            if all(v is None or str(v).strip() == "" for v in cell_values):
                continue

            record = {}
            for idx, db_col in col_map.items():
                if idx < len(cell_values):
                    val = cell_values[idx]
                    if db_col in _EURO_COLUMNS:
                        val = _parse_euro_value(val)
                    elif db_col in _DATE_COLUMNS:
                        val = _parse_date(val)
                    record[db_col] = val

            # Only insert if we have at least tipologia
            if not record.get("tipologia"):
                continue

            columns = ", ".join(record.keys())
            placeholders = ", ".join(["?"] * len(record))
            db.execute(
                f"INSERT INTO bfp ({columns}) VALUES ({placeholders})",
                list(record.values()),
            )
            count += 1

        db.commit()
        db.close()
        wb.close()
        return {"success": True, "count": count}
    except Exception as e:
        return {"error": str(e)}
