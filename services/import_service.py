import csv
import os

from database import get_db


# Mapping from Excel sheet names to database table names
SHEET_TABLE_MAP = {
    "ETF": "etf",
    "BOND": "bond",
    "Patrimonio": "patrimonio",
    "Investimento liquidita": "investimento_liquidita",
    "Simulazione_investimento": "simulazione_investimento",
}

# Mapping from Excel header names to DB column names
HEADER_MAP = {
    # ETF sheet
    "isin": "isin",
    "nome": "nome",
    "data di acquisto": "data_acquisto",
    "quantita": "quantita",
    "prezzo di acquisto": "prezzo_acquisto",
    "commissioni": "commissioni",
    "costo totale": "costo_totale",
    # BOND sheet
    "rateo lordo": "rateo_lordo",
    "ritenuta rateo": "ritenuta_rateo",
    "ritenuta dis.": "ritenuta_dis",
    "uscita": "uscita",
    "uscita ( x calcolo )": "uscita_per_calcolo",
    "quotazione eur": "quotazione_eur",
    "coeff. indicizz": "coeff_indicizz",
    # Patrimonio sheet
    "data": "data",
    "immobili esteri": "immobili_esteri",
    "immobile italia": "immobile_italia",
    "fondo pensione": "fondo_pensione",
    "etf": "etf",
    "bfp": "bfp",
    "btp": "btp",
    "cash": "cash",
    "cd": "cd",
    "tfr netto": "tfr_netto",
    "debiti": "debiti",
    # Investimento liquidita sheet
    "anno": "anno",
    "mese": "mese",
    "liquidita' in \nentrata": "liquidita_entrata",
    "liquidita' in entrata": "liquidita_entrata",
    "importo \nsu cd/xeon": "importo_cd",
    "importo su cd/xeon": "importo_cd",
    "importo \nsu bfpi": "importo_bfpi",
    "importo su bfpi": "importo_bfpi",
    "liquidita' \naccumulata": "liquidita_accumulata",
    "liquidita' accumulata": "liquidita_accumulata",
    "accumulo \nbsf": "accumulo_bsf",
    "accumulo bsf": "accumulo_bsf",
    "accumulo \nbfpi": "accumulo_bfpi",
    "accumulo bfpi": "accumulo_bfpi",
    "% bfpi": "pct_bfpi",
    "accumulo \ncd/xeon": "accumulo_cd",
    "accumulo cd/xeon": "accumulo_cd",
    "% cd/\nxeon": "pct_cd",
    "% cd/xeon": "pct_cd",
    "accumulo\naltro": "accumulo_altro",
    "accumulo altro": "accumulo_altro",
    "% altro": "pct_altro",
    "totale \nbsf": "totale_bsf",
    "totale bsf": "totale_bsf",
    "totale \naltro": "totale_altro",
    "totale altro": "totale_altro",
    "mesi passati": "mesi_passati",
    # Simulazione investimento sheet
    "immobili esteri": "immobili_esteri",
    "immobile italia": "immobile_italia",
    "fondo pensione": "fondo_pensione",
}


def import_from_excel(filepath):
    """Reads an Excel file using openpyxl and imports ALL sheets into their
    respective tables based on SHEET_TABLE_MAP.

    Returns dict with count of imported rows per table.
    """
    try:
        import openpyxl

        if not os.path.exists(filepath):
            return {"error": f"File non trovato: {filepath}"}

        wb = openpyxl.load_workbook(filepath, data_only=True)
        db = get_db()
        result = {}

        for sheet_name, table_name in SHEET_TABLE_MAP.items():
            if sheet_name not in wb.sheetnames:
                result[table_name] = 0
                continue

            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))

            if len(rows) < 2:
                result[table_name] = 0
                continue

            # First row is headers - map them to DB column names
            raw_headers = [str(h).strip().lower() if h else f"col_{i}" for i, h in enumerate(rows[0])]
            headers = [HEADER_MAP.get(h, h) for h in raw_headers]
            data_rows = rows[1:]

            # Get table column names from DB
            cursor = db.execute(f"PRAGMA table_info({table_name})")
            table_cols = {col[1] for col in cursor.fetchall()}

            # Filter headers to only those that exist in the table (excluding id)
            valid_indices = []
            valid_headers = []
            for i, h in enumerate(headers):
                if h in table_cols and h != "id":
                    valid_indices.append(i)
                    valid_headers.append(h)

            if not valid_headers:
                result[table_name] = 0
                continue

            columns = ", ".join(valid_headers)
            placeholders = ", ".join(["?"] * len(valid_headers))
            count = 0

            for row in data_rows:
                values = []
                skip = False
                for idx in valid_indices:
                    val = row[idx] if idx < len(row) else None
                    values.append(val)
                # Skip rows where all values are None
                if all(v is None for v in values):
                    continue
                try:
                    db.execute(
                        f"INSERT INTO {table_name} ({columns}) VALUES ({placeholders})",
                        values,
                    )
                    count += 1
                except Exception:
                    continue

            result[table_name] = count

        db.commit()
        db.close()
        wb.close()
        return result
    except Exception as e:
        return {"error": str(e)}


def export_to_csv(table_name, filepath):
    """Export a database table to a CSV file."""
    try:
        db = get_db()
        rows = db.execute(f"SELECT * FROM {table_name}").fetchall()
        db.close()

        if not rows:
            return {"error": f"Nessun dato nella tabella {table_name}"}

        headers = rows[0].keys()

        os.makedirs(os.path.dirname(filepath) or ".", exist_ok=True)
        with open(filepath, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f, delimiter=";")
            writer.writerow(headers)
            for row in rows:
                writer.writerow([row[h] for h in headers])

        return {"success": True, "rows": len(rows), "filepath": filepath}
    except Exception as e:
        return {"error": str(e)}


def import_from_csv(table_name, filepath):
    """Import data from a CSV file into a database table."""
    try:
        if not os.path.exists(filepath):
            return {"error": f"File non trovato: {filepath}"}

        db = get_db()

        # Get table column names
        cursor = db.execute(f"PRAGMA table_info({table_name})")
        table_cols = {col[1] for col in cursor.fetchall()}

        with open(filepath, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f, delimiter=";")

            # Filter to valid columns (excluding id)
            valid_headers = [
                h for h in reader.fieldnames if h in table_cols and h != "id"
            ]

            if not valid_headers:
                db.close()
                return {"error": "Nessuna colonna valida trovata nel CSV"}

            columns = ", ".join(valid_headers)
            placeholders = ", ".join(["?"] * len(valid_headers))
            count = 0

            for row in reader:
                values = [row.get(h) for h in valid_headers]
                if all(v is None or v == "" for v in values):
                    continue
                try:
                    db.execute(
                        f"INSERT INTO {table_name} ({columns}) VALUES ({placeholders})",
                        values,
                    )
                    count += 1
                except Exception:
                    continue

        db.commit()
        db.close()
        return {"success": True, "rows_imported": count}
    except Exception as e:
        return {"error": str(e)}


def export_to_excel(tables, filepath):
    """Export specified tables to Excel with formatting (headers bold, auto column width).

    Uses xlsxwriter.
    tables: list of table names to export.
    """
    try:
        import xlsxwriter

        db = get_db()
        os.makedirs(os.path.dirname(filepath) or ".", exist_ok=True)

        workbook = xlsxwriter.Workbook(filepath)
        bold_fmt = workbook.add_format({"bold": True, "bg_color": "#D9E1F2"})

        for table_name in tables:
            rows = db.execute(f"SELECT * FROM {table_name}").fetchall()
            if not rows:
                continue

            headers = rows[0].keys()
            worksheet = workbook.add_worksheet(table_name[:31])  # sheet name max 31 chars

            # Write headers
            for col, header in enumerate(headers):
                worksheet.write(0, col, header, bold_fmt)

            # Write data
            for row_idx, row in enumerate(rows, start=1):
                for col_idx, header in enumerate(headers):
                    value = row[header]
                    if value is not None:
                        worksheet.write(row_idx, col_idx, value)

            # Auto-fit column widths
            for col_idx, header in enumerate(headers):
                max_len = len(str(header))
                for row in rows:
                    cell_len = len(str(row[header])) if row[header] is not None else 0
                    max_len = max(max_len, cell_len)
                worksheet.set_column(col_idx, col_idx, min(max_len + 2, 50))

        workbook.close()
        db.close()
        return {"success": True, "filepath": filepath, "tables": tables}
    except Exception as e:
        return {"error": str(e)}


def export_to_excel_with_charts(filepath):
    """Export patrimonio data with embedded charts using xlsxwriter.

    Sheets:
        1. Patrimonio - Full data table with all asset classes, debiti, totals
        2. Variazioni - Percentage changes between consecutive records
        3. Allocazione - Latest allocation breakdown with pie chart
        4. Evoluzione - Stacked area chart + totale netto line
    """
    try:
        import xlsxwriter
        from services.patrimonio_service import (
            ASSET_FIELDS,
            get_all_patrimonio,
            get_patrimonio_percentuali,
            get_patrimonio_totali,
        )

        ASSET_LABELS = {
            "immobili_esteri": "Imm. Esteri",
            "immobile_italia": "Imm. Italia",
            "fondo_pensione": "Fondo Pens.",
            "etf": "ETF",
            "bfp": "BFP",
            "btp": "BTP",
            "cash": "Cash",
            "cd": "CD",
            "tfr_netto": "TFR Netto",
        }

        records = get_all_patrimonio()
        if isinstance(records, dict) and "error" in records:
            return records
        if not records:
            return {"error": "Nessun dato patrimonio disponibile"}

        os.makedirs(os.path.dirname(filepath) or ".", exist_ok=True)
        workbook = xlsxwriter.Workbook(filepath)

        # --- Formats ---
        bold_fmt = workbook.add_format({"bold": True, "bg_color": "#D9E1F2"})
        money_fmt = workbook.add_format({"num_format": "#,##0.00"})
        pct_fmt = workbook.add_format({"num_format": "0.00%"})
        pct_pos = workbook.add_format({"num_format": "0.00%", "font_color": "#008000"})
        pct_neg = workbook.add_format({"num_format": "0.00%", "font_color": "#CC0000"})
        bold_money = workbook.add_format({"bold": True, "num_format": "#,##0.00"})

        # =============================================
        # SHEET 1: Patrimonio (data table)
        # =============================================
        ws_data = workbook.add_worksheet("Patrimonio")
        headers = ["Data"] + [ASSET_LABELS.get(f, f) for f in ASSET_FIELDS] + ["Debiti", "Totale Netto"]
        for col, header in enumerate(headers):
            ws_data.write(0, col, header, bold_fmt)

        totali_list = []
        for row_idx, record in enumerate(records, start=1):
            ws_data.write(row_idx, 0, record.get("data", ""))
            for col_idx, field in enumerate(ASSET_FIELDS, start=1):
                ws_data.write(row_idx, col_idx, float(record.get(field, 0) or 0), money_fmt)
            ws_data.write(row_idx, len(ASSET_FIELDS) + 1, float(record.get("debiti", 0) or 0), money_fmt)
            totali = get_patrimonio_totali(record)
            netto = totali.get("totale_netto", 0) if not isinstance(totali, dict) or "error" not in totali else 0
            totali_list.append(netto)
            ws_data.write(row_idx, len(ASSET_FIELDS) + 2, netto, bold_money)

        ws_data.set_column(0, 0, 12)
        ws_data.set_column(1, len(headers) - 1, 15)
        num_rows = len(records)

        # =============================================
        # SHEET 2: Variazioni %
        # =============================================
        ws_var = workbook.add_worksheet("Variazioni")
        var_headers = ["Data"] + [ASSET_LABELS.get(f, f) for f in ASSET_FIELDS] + ["Totale Netto"]
        for col, header in enumerate(var_headers):
            ws_var.write(0, col, header, bold_fmt)

        for row_idx in range(1, len(records)):
            curr = records[row_idx]
            prev = records[row_idx - 1]
            ws_var.write(row_idx, 0, curr.get("data", ""))
            for col_idx, field in enumerate(ASSET_FIELDS, start=1):
                cv = float(curr.get(field, 0) or 0)
                pv = float(prev.get(field, 0) or 0)
                if pv != 0:
                    var = (cv - pv) / abs(pv)
                    fmt = pct_pos if var >= 0 else pct_neg
                    ws_var.write(row_idx, col_idx, var, fmt)
                else:
                    ws_var.write(row_idx, col_idx, "")
            # Totale netto variation
            cn = totali_list[row_idx]
            pn = totali_list[row_idx - 1]
            if pn != 0:
                var = (cn - pn) / abs(pn)
                fmt = pct_pos if var >= 0 else pct_neg
                ws_var.write(row_idx, len(ASSET_FIELDS) + 1, var, fmt)

        ws_var.set_column(0, 0, 12)
        ws_var.set_column(1, len(var_headers) - 1, 15)

        # =============================================
        # SHEET 3: Allocazione (pie chart)
        # =============================================
        latest = records[-1]
        percentuali = get_patrimonio_percentuali(latest)

        ws_alloc = workbook.add_worksheet("Allocazione")
        ws_alloc.write(0, 0, "Asset Class", bold_fmt)
        ws_alloc.write(0, 1, "Valore", bold_fmt)
        ws_alloc.write(0, 2, "Percentuale", bold_fmt)

        colors = ['#4472C4', '#ED7D31', '#A5A5A5', '#FFC000', '#70AD47',
                  '#5B9BD5', '#FF6384', '#9966FF', '#FF9F40']
        points = []
        for i, field in enumerate(ASSET_FIELDS):
            ws_alloc.write(i + 1, 0, ASSET_LABELS.get(field, field))
            ws_alloc.write(i + 1, 1, float(latest.get(field, 0) or 0), money_fmt)
            pct = percentuali.get(field, 0) if isinstance(percentuali, dict) else 0
            ws_alloc.write(i + 1, 2, pct / 100 if pct else 0, pct_fmt)
            if i < len(colors):
                points.append({"fill": {"color": colors[i]}})

        ws_alloc.set_column(0, 0, 18)
        ws_alloc.set_column(1, 2, 15)

        pie_chart = workbook.add_chart({"type": "pie"})
        pie_chart.add_series({
            "name": "Allocazione Patrimonio",
            "categories": ["Allocazione", 1, 0, len(ASSET_FIELDS), 0],
            "values": ["Allocazione", 1, 1, len(ASSET_FIELDS), 1],
            "data_labels": {"percentage": True, "position": "inside_end",
                            "font": {"size": 10, "bold": True, "color": "white"}},
            "points": points,
        })
        pie_chart.set_title({"name": "Allocazione Patrimonio"})
        pie_chart.set_legend({"position": "bottom", "font": {"size": 10}})
        pie_chart.set_size({"width": 720, "height": 500})
        ws_alloc.insert_chart("E2", pie_chart)

        # =============================================
        # SHEET 4: Evoluzione (stacked area + line)
        # =============================================
        ws_evo = workbook.add_worksheet("Evoluzione")

        # Stacked area chart
        area_chart = workbook.add_chart({"type": "area", "subtype": "stacked"})
        for i, field in enumerate(ASSET_FIELDS):
            col_idx = i + 1
            area_chart.add_series({
                "name": ASSET_LABELS.get(field, field),
                "categories": ["Patrimonio", 1, 0, num_rows, 0],
                "values": ["Patrimonio", 1, col_idx, num_rows, col_idx],
                "fill": {"color": colors[i] if i < len(colors) else "#888888"},
                "border": {"none": True},
            })
        area_chart.set_title({"name": "Evoluzione Patrimonio per Asset Class"})
        area_chart.set_x_axis({"name": "Data"})
        area_chart.set_y_axis({"name": "EUR", "num_format": "#,##0"})
        area_chart.set_size({"width": 900, "height": 500})
        area_chart.set_legend({"position": "bottom", "font": {"size": 9}})
        ws_evo.insert_chart("A1", area_chart)

        # Totale netto line chart
        line_chart = workbook.add_chart({"type": "line"})
        line_chart.add_series({
            "name": "Totale Netto",
            "categories": ["Patrimonio", 1, 0, num_rows, 0],
            "values": ["Patrimonio", 1, len(ASSET_FIELDS) + 2, num_rows, len(ASSET_FIELDS) + 2],
            "line": {"width": 2.5, "color": "#4472C4"},
            "marker": {"type": "circle", "size": 4},
        })
        line_chart.set_title({"name": "Evoluzione Totale Netto"})
        line_chart.set_x_axis({"name": "Data"})
        line_chart.set_y_axis({"name": "EUR", "num_format": "#,##0"})
        line_chart.set_size({"width": 900, "height": 400})
        ws_evo.insert_chart("A26", line_chart)

        workbook.close()
        return {"success": True, "filepath": filepath}
    except Exception as e:
        return {"error": str(e)}
